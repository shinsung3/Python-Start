from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
import openpyxl

#driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver = webdriver.Chrome('./chromedriver')
driver.get('https://www.elevator.go.kr/main/Login.do?menuId=00000000')


####################################################
Month = '01'
YMD = '20220303'
####################################################


# 국가승강기센터 아이디 모음
userId = ["helco2024063", "helco7840", "chanwool2", "kkongs2k", "h2026236", "helco5005655", "helco5013119", "helco3016471",
          "kun072222", "helco2024209", "hdel2014882", "helco5007216", "hanbit09", "helco202413", "helco2017130", "helco2024683", "hds2900513"]
password = ["helco0318@", "J2027011!", "1268104267q!", "kk7270!!kk", "wjsgkfkd1!", "akfkdud123!", "mj5013119*", "rldnjs*001",
            "3009181@kdk", "skywjd35^", "helco2014882*", "guseo1234!", "1023201aA@@", "!helco2024136", "helco200++", "sjh159951*", "hds2900513!!"]

####################################################
for item in range(0, 16):
    userId[item]
    password[item]

# 다음 세 줄이 기본 패턴 코드: ID 넣기
# WebDriverWait(driver, 최대 기다리는 시간).until(EC.presence_of_element_located((By.CSS_SELECTOR, CSS Selector 태그)))
    login_id = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "#userId")))
    login_id.clear()  # 입력창의 경우, 사전에 작성되어 있는 텍스트를 삭제
    login_id.send_keys(userId[item])  # 내가 넣고자 하는 텍스트 삽입

# 다음 세 줄이 기본 패턴 코드: 패스워드 넣기
    login_pw = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "#password")))
    login_pw.clear()
    login_pw.send_keys(password[item])

# 버튼 클릭시는 다음 두 줄: 로그인 버튼 누르기
    button = WebDriverWait(driver, 3).until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "#sb_contents > div > div.login_box2 > fieldset > form:nth-child(2) > div > a")))
    button.click()
    time.sleep(1.5)  # 로그인 후의 페이지 로딩을 위해, 1초정도 기다리면 좋음

    driver.get('https://www.elevator.go.kr/rep/selfCheckCmplL.do?menuId=IW170807')

# 연도 선택 2022
    button = WebDriverWait(driver, 3).until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "#search_y > option:nth-child(12)")))
    button.click()
    time.sleep(1.5)  # 로그인 후의 페이지 로딩을 위해, 1초정도 기다리면 좋음

# 월 선택 1월
    button = WebDriverWait(driver, 3).until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "#search_m > option:nth-child(2)")))
    button.click()
    time.sleep(1.5)  # 로그인 후의 페이지 로딩을 위해, 1초정도 기다리면 좋음

# 검색 버튼 클릭
    button = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "#button > div.sy_3 > a")))
    button.click()
    time.sleep(5)  # 다운로드 후 5초 기다리기

# 엑셀 다운로드 버튼
    button = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "#button > div.sy_5 > a")))
    button.click()
    time.sleep(4)  # 다운로드 후 5초 기다리기

# 로그아웃
    button = WebDriverWait(driver, 3).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "#welcome > ul > li.logout > a")))
    button.click()
    time.sleep(1)  # 로그아웃

####################################################
userId[16]
password[16]

# 다음 세 줄이 기본 패턴 코드: ID 넣기
# WebDriverWait(driver, 최대 기다리는 시간).until(EC.presence_of_element_located((By.CSS_SELECTOR, CSS Selector 태그)))
login_id = WebDriverWait(driver, 3).until(
    EC.presence_of_element_located((By.CSS_SELECTOR, "#userId")))
login_id.clear()  # 입력창의 경우, 사전에 작성되어 있는 텍스트를 삭제
login_id.send_keys(userId[16])  # 내가 넣고자 하는 텍스트 삽입

# 다음 세 줄이 기본 패턴 코드: 패스워드 넣기
login_pw = WebDriverWait(driver, 3).until(
    EC.presence_of_element_located((By.CSS_SELECTOR, "#password")))
login_pw.clear()
login_pw.send_keys(password[16])

# 버튼 클릭시는 다음 두 줄: 로그인 버튼 누르기
button = WebDriverWait(driver, 3).until(EC.presence_of_element_located(
    (By.CSS_SELECTOR, "#sb_contents > div > div.login_box2 > fieldset > form:nth-child(2) > div > a")))
button.click()
time.sleep(1.5)  # 로그인 후의 페이지 로딩을 위해, 1초정도 기다리면 좋음

driver.get('https://www.elevator.go.kr/rep/selfCheckCmplL.do?menuId=IW170807')

# 연도 선택 2022
button = WebDriverWait(driver, 3).until(EC.presence_of_element_located(
    (By.CSS_SELECTOR, "#search_y > option:nth-child(12)")))
button.click()
time.sleep(1.5)  # 로그인 후의 페이지 로딩을 위해, 1초정도 기다리면 좋음

# 월 선택 1월
button = WebDriverWait(driver, 3).until(EC.presence_of_element_located(
    (By.CSS_SELECTOR, "#search_m > option:nth-child(2)")))
button.click()
time.sleep(1.5)  # 로그인 후의 페이지 로딩을 위해, 1초정도 기다리면 좋음

# 검색 버튼 클릭
button = WebDriverWait(driver, 3).until(
    EC.presence_of_element_located((By.CSS_SELECTOR, "#button > div.sy_3 > a")))
button.click()
time.sleep(5)  # 다운로드 후 5초 기다리기

# 엑셀 다운로드 버튼
button = WebDriverWait(driver, 3).until(
    EC.presence_of_element_located((By.CSS_SELECTOR, "#button > div.sy_5 > a")))
button.click()
time.sleep(10)  # 다운로드 후 10초 기다리기

# 로그아웃
button = WebDriverWait(driver, 3).until(
    EC.presence_of_element_located((By.CSS_SELECTOR, "#welcome > ul > li.logout > a")))
button.click()
time.sleep(1)  # 로그아웃
####################################################

driver.quit()

# 엑셀 합치기
fPath = f'C:\\Users\\Administrator\\Downloads\\2022년+{Month}월+자체점검실시내역{YMD}'

Name_A = f"2022년+{Month}월+자체점검실시내역{YMD}"
Name_B = f"2022년+{Month}월+자체점검실시내역{YMD} (1)"
Name_C = f"2022년+{Month}월+자체점검실시내역{YMD} (2)"
Name_D = f"2022년+{Month}월+자체점검실시내역{YMD} (3)"
Name_E = f"2022년+{Month}월+자체점검실시내역{YMD} (4)"
Name_F = f"2022년+{Month}월+자체점검실시내역{YMD} (5)"
Name_G = f"2022년+{Month}월+자체점검실시내역{YMD} (6)"
Name_H = f"2022년+{Month}월+자체점검실시내역{YMD} (7)"
Name_I = f"2022년+{Month}월+자체점검실시내역{YMD} (8)"
Name_J = f"2022년+{Month}월+자체점검실시내역{YMD} (9)"
Name_K = f"2022년+{Month}월+자체점검실시내역{YMD} (10)"
Name_L = f"2022년+{Month}월+자체점검실시내역{YMD} (11)"
Name_M = f"2022년+{Month}월+자체점검실시내역{YMD} (12)"
Name_N = f"2022년+{Month}월+자체점검실시내역{YMD} (13)"
Name_O = f"2022년+{Month}월+자체점검실시내역{YMD} (14)"
Name_P = f"2022년+{Month}월+자체점검실시내역{YMD} (15)"
Name_Q = f"2022년+{Month}월+자체점검실시내역{YMD} (16)"

Name_A = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_A}.xls')
NewName_A = Name_A.drop([0, 1, 2, 3])

Name_B = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_B}.xls')
NewName_B = Name_B.drop([0, 1, 2, 3])

Name_C = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_C}.xls')
NewName_C = Name_C.drop([0, 1, 2, 3])

Name_D = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_D}.xls')
NewName_D = Name_D.drop([0, 1, 2, 3])

Name_E = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_E}.xls')
NewName_E = Name_E.drop([0, 1, 2, 3])

Name_F = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_F}.xls')
NewName_F = Name_F.drop([0, 1, 2, 3])

Name_G = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_G}.xls')
NewName_G = Name_G.drop([0, 1, 2, 3])

Name_H = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_H}.xls')
NewName_H = Name_H.drop([0, 1, 2, 3])

Name_I = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_I}.xls')
NewName_I = Name_I.drop([0, 1, 2, 3])

Name_J = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_J}.xls')
NewName_J = Name_J.drop([0, 1, 2, 3])

Name_K = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_K}.xls')
NewName_K = Name_K.drop([0, 1, 2, 3])

Name_L = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_L}.xls')
NewName_L = Name_L.drop([0, 1, 2, 3])

Name_M = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_M}.xls')
NewName_M = Name_M.drop([0, 1, 2, 3])

Name_N = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_N}.xls')
NewName_N = Name_N.drop([0, 1, 2, 3])

Name_O = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_O}.xls')
NewName_O = Name_O.drop([0, 1, 2, 3])

Name_P = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_P}.xls')
NewName_P = Name_P.drop([0, 1, 2, 3])

Name_Q = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_Q}.xls')
NewName_Q = Name_Q.drop([0, 1, 2, 3])

Name_concat = pd.concat([NewName_A, NewName_B, NewName_C, NewName_D, NewName_E, NewName_F, NewName_G, NewName_H,
                         NewName_I, NewName_J, NewName_K, NewName_L, NewName_M, NewName_N, NewName_O, NewName_P, NewName_Q], axis=0)
Name_concat.to_excel(
    f'C:\\Users\\Administrator\\Downloads\\2022년+{Month}월+자체점검실시내역{YMD}_통합.xlsx')

# 샘플 3개
# Name_concat = pd.concat([NewName_A, NewName_B, NewName_C], axis=0)
# Name_concat.to_excel(
#     f'C:\\Users\\Administrator\\Downloads\\2022년+{Month}월+자체점검실시내역{YMD}_통합.xlsx')

# 첫 번째 행 수정
wb = openpyxl.load_workbook(
    f'C:\\Users\\Administrator\\Downloads\\2022년+{Month}월+자체점검실시내역{YMD}_통합.xlsx')

ws = wb.active

ws['A1'] = 'No'
ws['B1'] = '점검업체'
ws['C1'] = '점검일자'
ws['D1'] = '시작시간'
ws['E1'] = '종료시간'
ws['F1'] = '전산등록일'
ws['G1'] = '전산수정일'
ws['H1'] = '건물명'
ws['I1'] = '건물주소1'
ws['J1'] = '건물주소2'
ws['K1'] = '승강기고유번호'
ws['L1'] = '호기'
ws['M1'] = '설치장소'
ws['N1'] = '승강기종류'
ws['O1'] = '점검자'
ws['P1'] = '점검보조자'
ws['Q1'] = '관리주체확인여부'
ws['R1'] = '관리주체확인일자'
ws['S1'] = '점검결과'

wb.save(
    f'C:\\Users\\Administrator\\Downloads\\2022년+{Month}월+자체점검실시내역{YMD}_통합.xlsx')


# 년도
# search_y > option:nth-child(11)

# 1월
# search_m > option:nth-child(2)
# 2월
# search_m > option:nth-child(3)
# 3월
# search_m > option:nth-child(4)
# 4월
# search_m > option:nth-child(5)
# 5월
# search_m > option:nth-child(6)
# 6월
# search_m > option:nth-child(7)
# 7월
# search_m > option:nth-child(8)
# 8월
# search_m > option:nth-child(9)
# 9월
# search_m > option:nth-child(10)
# 0월
# search_m > option:nth-child(11)
# 11월
# search_m > option:nth-child(12)
# 12월
# search_m > option:nth-child(13)
